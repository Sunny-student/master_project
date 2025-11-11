import csv
import openpyxl
import os

# 要寫入的欄位內容（A1 到 O1）
headers = [
    "Serial_Number", "5000-6900", "5000-6909", "5000-6906", "5000-6941", "5000-6900N",
    "5000-6937", "5000-6912", "5000-6926", "5000-6930", "5000-6903",
    "5000-6911", "5000-6916", "5000-6905", "5000-6960", "5000-6914",
    "Target_luster", "Target_L", "Target_a", "Target_b" 
]

serial_5000 = [
    "5000-6900", "5000-6909", "5000-6906", "5000-6941", "5000-6900N",
    "5000-6937", "5000-6912", "5000-6926", "5000-6930", "5000-6903",
    "5000-6911", "5000-6916", "5000-6905", "5000-6960", "5000-6914"]

csv_path = "油漆調色_原始資料/painting_data.csv"
# 建立 CSV 檔案
with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
    writer = csv.writer(f)
    writer.writerow(headers)

excel_path = "C:/Users/robotic/Desktop/Luke/油漆調色_原始資料/251016/112年-114年測量數據(以合併).xlsx"
wb = openpyxl.load_workbook(excel_path, data_only=True)
sheet = wb.active


for row in range(2, sheet.max_row + 1):
    d = sheet[f"D{row}"].value
    if isinstance(d, (int, float)):

        # 取得 C~G 的內容
        c = sheet[f"C{row}"].value
        d = sheet[f"D{row}"].value
        e = sheet[f"E{row}"].value
        f = sheet[f"F{row}"].value
        g = sheet[f"G{row}"].value

        # --- Step 3: 將資料寫入 CSV 的 A2、Q2~T2 ---
        # A2 → Serial_Number 對應 C2
        # Q2 → Target_L 對應 D2
        # R2 → Target_a 對應 E2
        # S2 → Target_b 對應 F2
        # T2 → Target_luster 對應 G2

        row_data = [c] + [""] * 15 + [d, e, f, g]  # 共 20 欄



        # print("✅ 已建立 painting_data.csv 並成功寫入 A2、Q2~T2 的內容。")

        # === Step 3: 在 My Documents 中搜尋名為 C2 的 Excel 檔 ===
        documents_path = "C:/Users/robotic/Desktop/Luke/油漆調色_原始資料/My Documents"  # My Documents 的路徑
        target_filename = f"{c}.xlsx"

        found_path = None
        for root, dirs, files in os.walk(documents_path):
            if target_filename in files:
                found_path = os.path.join(root, target_filename)
                break

        if found_path:
            wb_target = openpyxl.load_workbook(found_path, data_only=True)
            sheet_target = wb_target.active

            for row2 in range(2, sheet_target.max_row + 1):  # 從第2列開始
                j_value = sheet_target[f"J{row2}"].value
                d_value = sheet_target[f"D{row2}"].value
                # 若 J 欄有數值且不為 0
                if isinstance(j_value, (int, float)) and j_value != 0:
                    # 若 D 欄內容尚未印過
                    if d_value.startswith("5000"):
                        if (d_value[-2:] == "停產") or (d_value[-2:] == "停用"):
                            d_value = d_value[:-2]
                            print(f"{found_path}的{d_value}已刪除'停產、停用'")

                        if d_value in serial_5000:
                            row_data[headers.index(d_value)] = j_value
                            
                        else:
                            print(f"資料整理有誤：{d_value}沒有記在標頭裡")
            
            row_data = [0 if x == "" else x for x in row_data]

            with open(csv_path, "a", newline="", encoding="utf-8-sig") as f:
                                writer = csv.writer(f)
                                writer.writerow(row_data)

                        # print(printed)
            # j3_value = sheet_target["J3"].value
            # print(f"📄 在 {found_path} 找到檔案，J3 的內容是：{j3_value}")
        else:
            print(f"❌ 在 {documents_path} 未找到 {target_filename}")

