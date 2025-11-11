# import openpyxl

# # 指定 Excel 檔案路徑
# file_path = "C:/Users/robotic/Desktop/Luke/油漆調色_原始資料/251016/114年1-9月調色.xlsx"

# # 讀取檔案
# wb = openpyxl.load_workbook(file_path, data_only=True)
# sheet = wb.active

# # 取得 V5 的內容
# cell_value = str(sheet["V5"].value)

# # 檢查是否完全等於指定字串（包含尾端空白）
# if cell_value == "P05000A01C25G30":
#     print("✅ V5 的內容正確！")
# else:
#     print(f"❌ V5 內容不符，實際為：'{cell_value}'")


# import openpyxl

# # 指定 Excel 檔案路徑
# file_path = "C:/Users/robotic/Desktop/Luke/油漆調色_原始資料/251016/112年-114年測量數據(以合併).xlsx"

# # 讀取 Excel
# wb = openpyxl.load_workbook(file_path, data_only=True)
# sheet = wb.active

# # 統計 E 欄中（從第2列開始）非空白的儲存格數
# count = 0
# for cell in sheet["B"][1:]:  # [1:] 是跳過 row 1
#     if isinstance(cell.value, (int, float)):  # 檢查是否為數值型態
#         count += 1

# print(f"✅ E 欄（從第2列起）共有 {count} 個有內容的儲存格。")


# import os

# import openpyxl

# # 指定 Excel 檔案路徑
# folder_path = "C:/Users/robotic/Desktop/Luke/油漆調色_原始資料/My Documents"
# # file_path = "C:/Users/robotic/Desktop/Luke/油漆調色_原始資料/My Documents/GMPC05000C23A07.xlsx"

# printed = []  # 用來記錄已經印過的 D 欄內容

# file_path = "C:/Users/robotic/Desktop/Luke/油漆調色_原始資料/My Documents/GMPC05000C23A07.xlsx"
# wb = openpyxl.load_workbook(file_path, data_only=True)
# sheet = wb.active


# for row in range(2, sheet.max_row + 1):  # 從第2列開始
#     j_value = sheet[f"J{row}"].value
#     d_value = sheet[f"D{row}"].value

#     # 若 J 欄有數值且不為 0
#     if isinstance(j_value, (int, float)) and j_value != 0:
#         # 若 D 欄內容尚未印過
#         if d_value not in printed:
#             printed.append(d_value)
#             print(d_value)
#             # print(printed)

# for filename in os.listdir(folder_path):
#     if filename.lower().endswith(".xlsx"):
#         file_path = os.path.join(folder_path, filename)
#         # print(file_path)
#         # 讀取 Excel
#         wb = openpyxl.load_workbook(file_path, data_only=True)
#         sheet = wb.active


#         for row in range(2, sheet.max_row + 1):  # 從第2列開始
#             j_value = sheet[f"J{row}"].value
#             d_value = sheet[f"D{row}"].value

#             # 若 J 欄有數值且不為 0
#             if isinstance(j_value, (int, float)) and j_value != 0:
#                 # 若 D 欄內容尚未印過
#                 if d_value not in printed:
#                     printed.append(d_value)
#                     print(d_value)
#                     # if not d_value.startswith("5000"):
#                         # print(file_path)
#                     # print(file_path)

# print(len(printed))


import os

import openpyxl

# 指定 Excel 檔案路徑
folder_path = "C:/Users/robotic/Desktop/Luke/油漆調色_原始資料/My Documents"
# file_path = "C:/Users/robotic/Desktop/Luke/油漆調色_原始資料/My Documents/GMPC05000C23A07.xlsx"
count = 0
printed = []  # 用來記錄已經印過的 D 欄內容

file_path = "C:/Users/robotic/Desktop/Luke/油漆調色_原始資料/My Documents/GMPC05000C23A07.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)
sheet = wb.active


for row in range(2, sheet.max_row + 1):  # 從第2列開始
    j_value = sheet[f"J{row}"].value
    d_value = sheet[f"D{row}"].value

    # 若 J 欄有數值且不為 0
    if isinstance(j_value, (int, float)) and j_value != 0:
        # 若 D 欄內容尚未印過
        if d_value.startswith("P8"):
            count+=1
            printed.append(d_value)
            print(d_value)
            print(file_path)
            # print(printed)

for filename in os.listdir(folder_path):
    if filename.lower().endswith(".xlsx"):
        file_path = os.path.join(folder_path, filename)
        # print(file_path)
        # 讀取 Excel
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb.active


        for row in range(2, sheet.max_row + 1):  # 從第2列開始
            j_value = sheet[f"J{row}"].value
            d_value = sheet[f"D{row}"].value

            # 若 J 欄有數值且不為 0
            if isinstance(j_value, (int, float)) and j_value != 0:
                # 若 D 欄內容尚未印過
                if d_value.startswith("P8"):
                    count+=1
                    printed.append(d_value)
                    print(d_value)
                    # if not d_value.startswith("5000"):
                        # print(file_path)
                    print(file_path)

print(len(printed))